import datetime
import pandas as pd
from openpyxl.styles import PatternFill
import numpy as np
from dados import Dados
import tkinter as tk
import pandas as pd
import win32com.client as win32
from tkinter import filedialog
from openpyxl import load_workbook
from functools import reduce
global arquivo_importado
arquivo_importado = None

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        x = self.widget.winfo_rootx() + self.widget.winfo_width()
        y = self.widget.winfo_rooty()
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip, text=self.text, background="lightyellow", relief="solid")
        label.pack()

    def hide_tooltip(self, event):
        if self.tooltip is not None:
            self.tooltip.destroy()
            self.tooltip = None

def importar_arquivo():
    global arquivo_importado
    arquivo = filedialog.askopenfilename(filetypes=[("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")])
    if arquivo:
        textbox.delete(0, tk.END)
        textbox.insert(tk.END, arquivo)
        arquivo_importado = arquivo

def conciliacao():
    day  = Dados().daycoval_data(arquivo_importado)
    data = datetime.date.today().strftime("%d_%m_%Y")
    path = f"historico\\CONCI_DAYC_ATV_{data}.xlsx"
    
    texto = "\nAcessando a Sinqia... Aguarde."
    tela_preta.insert(tk.END, texto)
    tela_preta.update_idletasks()
    
    atv = Dados().ativa_data()
    
    texto = "\n\nArquivo 'Posições Cotistas' (45) baixado. \nConciliação iniciada."
    tela_preta.insert(tk.END, texto)
    tela_preta.update_idletasks()
    
    
    # Realizando a fusão dos DataFrames com base nas colunas em comum
    merged_df = atv.merge(day, on=['Cliente',  'CNPJ_Fundo'], how='outer')

    # Calculando a diferença e o módulo da diferença
    merged_df['Diferença'] = abs(merged_df['Posição_Ativa'] - merged_df['Posição_Dayc'])

    # Incluindo o nome do fundo quando uma celula referente for vazia
    merged_df = merged_df.sort_values(by='Diferença', ascending=False)
    merged_df['Nome_Fundo_x'].replace('', np.nan, inplace=True)
    merged_df['Nome_Fundo_x'].fillna(merged_df['Nome_Fundo_y'], inplace=True)
    merged_df['Nome_Fundo_y'].replace('', np.nan, inplace=True)
    merged_df['Nome_Fundo_y'].fillna(merged_df['Nome_Fundo_x'], inplace=True)

    del merged_df['Nome_Fundo_y']

    merged_df.rename(columns={
                        'Nome_Fundo_x': 'Nome_Fundo'
                    }, inplace=True)
    #formatando para que não saia com notação científica no excel
    merged_df['Diferença'] = merged_df['Diferença'].apply(lambda x: f'{x:.2f}' if not pd.isnull(x) else 'nan')

    # Cria a coluna 'Posição_Superior' considerando as possibilidades de maior, menor ou igual
    conditions = [
        merged_df['Posição_Ativa'] > merged_df['Posição_Dayc'],
        merged_df['Posição_Ativa'] < merged_df['Posição_Dayc'],
        abs(merged_df['Posição_Ativa'] - merged_df['Posição_Dayc']) < 0.01
    ]
    choices = [
        'Posição_Ativa', 'Posição_Dayc', 'Iguais'
    ]
    merged_df['Posição_Superior'] = np.select(conditions, choices, default='')

    # Atualiza a coluna 'Posição_Superior' para ficar vazia se a diferença for menor que 0.01
    merged_df['Diferença'] = merged_df['Diferença'].astype(float)
    merged_df['Posição_Ativa'] = merged_df['Posição_Ativa'].astype(float)
    merged_df['Posição_Dayc'] = merged_df['Posição_Dayc'].astype(float)

    merged_df.loc[merged_df['Diferença'] < 0.01, 'Posição_Superior'] = ''

    merged_df['Posição_Ativa'] = merged_df['Posição_Ativa'].apply(lambda x: f'R${x:,.2f}')
    merged_df['Posição_Dayc'] = merged_df['Posição_Dayc'].apply(lambda x: f'R${x:,.2f}')

    # texto = "\nSalvando os dados conciliados..."
    # tela_preta.insert(tk.END, texto)
    # tela_preta.update_idletasks()

    # salvando o resultado
    merged_df.to_excel(path, index=False)

    # texto = "\nEditando a planilha..."
    # tela_preta.insert(tk.END, texto)
    # tela_preta.update_idletasks()
    
    # Obtém a folha do Excel e o formato de estilo
    writer = pd.ExcelWriter(path, engine='openpyxl')
    merged_df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Obtém a planilha
    workbook = writer.book
    worksheet = workbook.active

    # Define o estilo de célula amarela
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Itera pelas linhas e aplica o estilo amarelo às células com diferença maior que 100
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, min_col=merged_df.columns.get_loc('Cliente') + 1), start=2):
        if merged_df.iloc[row_idx - 2]['Diferença'] > 100:
            for cell in row:
                cell.fill = yellow_fill

    # Salva o arquivo Excel
    workbook.save(path)
    workbook.close()
    writer.close()
    
    texto = f"\n\nConcluído.\nPlanilha salva em {path}."
    tela_preta.insert(tk.END, texto)
    tela_preta.update_idletasks()
    
    
    

cor = 'DarkCyan'
janela = tk.Tk()
janela.title("DAYCONCI")
janela.geometry("500x400")
janela.resizable(False, False)
janela.configure(bg=cor)

# Criar frame para centralizar os elementos
frame = tk.Frame(janela)
frame.pack(pady=40)
frame.configure(bg=cor)

# Criar caixa de texto
textbox = tk.Entry(frame, width=50)
textbox.pack(side=tk.LEFT, padx=10, pady=10)

# Criar botão de importar
botao_importar = tk.Button(frame, text="Importar", command=importar_arquivo)
botao_importar.pack(side=tk.LEFT, pady=10)

# Criar frame para o botão processar
frame_processar = tk.Frame(janela)
frame_processar.pack(pady=20)
frame_processar.configure(bg=cor)

# Adicionar o símbolo de "?" ao lado do botão
simbolo_interrogacao = tk.Label(frame_processar, text="?", font=("Arial", 12, "bold"), fg="Aquamarine", cursor="question_arrow", bg=cor)
simbolo_interrogacao.pack(side=tk.RIGHT, padx=5)
ToolTip(simbolo_interrogacao, "Ao clicar em processar, será iniciada a conciliação com \nbase no arquivo importado (planilha Daycoval) e os dados da base que se refere ao 45. Ao término da conciliação, \nserá salvo o arquivo na pasta 'historico' neste mesmo diretório.")

# Criar botão de processar
botao_processar = tk.Button(frame_processar, text="Processar", command=conciliacao)
botao_processar.pack(side=tk.TOP, pady=10)

# Criar "tela preta" para as mensagens
tela_preta = tk.Text(janela, bg="black", fg="white", height=10)
tela_preta.pack(fill=tk.BOTH, padx=10, pady=10)


janela.mainloop()

